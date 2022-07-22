import csv
import openpyxl
from modules.helpers.get_headers_helper import get_headers
from modules.helpers.get_index_helper import get_index
from modules.helpers.get_csv_data_site_helper import get_csv_data_site
from modules.helpers.get_csv_incidents_counters_helper import get_csv_incidents_counters

def site_and_category_filler_by_week(quip_week_csv, sites_and_categories_csv, whs_anual_metrics_xlsx, incidents_count_and_rates_csv):
    """This function read .csv files from a directory, uses other .csv files as parameters and writes in to a existing .xlsx file"""

    # Load sites and categories from Sites and categories.csv and formule a dictionary to identify anc calssify wich category is each site from quip
    with open(sites_and_categories_csv) as f:
        reader = csv.reader(f, delimiter = ';')
        header_row = next(reader)

        list_sites_categories = []
        for row in reader:
            if row != []:
                site_category = {f'{header_row[0]}': f'{row[0]}', f'{header_row[1]}': f'{row[1]}'}
                list_sites_categories.append(site_category)
            else:
                continue

    # Load site data from Incidents count and rates.csv to formule 2 lists with keys as variables to work with
    # A list with counters incidents and a list with all data site
    list_incidents_counters = get_csv_incidents_counters(incidents_count_and_rates_csv)
    list_site_data = get_csv_data_site(incidents_count_and_rates_csv)

    # Start openpyxl library to open existing .xlsx file
    workbook = openpyxl.load_workbook(whs_anual_metrics_xlsx)

    # Opens Site metrics by week worksheet and Category metrics by week worksheet
    worksheet_site_metrics_by_week = workbook['Site metrics by week']
    worksheet_category_metrics_by_week = workbook['Category metrics by week']

    # Get 'Site metrics by week' worksheet headers
    worksheet_site_metrics_by_week_headers = get_headers(worksheet_site_metrics_by_week['1'])

    # Get 'Site metrics by week' worksheet sites names
    worksheet_site_metrics_by_week_sites_names = get_headers(worksheet_site_metrics_by_week['B'])

    # Get 'Category metrics by week' worksheet headers
    worksheet_category_metrics_by_week_headers = get_headers(worksheet_category_metrics_by_week['1'])

    # Get 'Category metrics by week' worksheet sites categories
    worksheet_category_metrics_by_week_categories_names = get_headers(worksheet_category_metrics_by_week['A'])

    # Creates a list of categories empty dictionaries from 'Category metrics by week' worksheet 
    # And fill each one with its incidents
    list_site_categories_wbr = set(worksheet_category_metrics_by_week_categories_names)
    list_site_categories_dictionaries_wbr = []
    for site_category_wbr in list_site_categories_wbr:
        site_category_dictionary_wbr = {
            'category':f'{site_category_wbr.lower()}', 
            'week': '', 
            'worked_hours': 0.00
            }
        for incident in list_incidents_counters:
            site_category_dictionary_wbr[incident] = 0
        list_site_categories_dictionaries_wbr.append(site_category_dictionary_wbr)

    # Creates a list of dictionaries of each site with its complety data of that week
    # Load quip information from 'quip.csv' to create a list of site dictioanries with data from that week
    with open(quip_week_csv) as f:
        reader = csv.reader(f, delimiter = ';')
        header_row = next(reader)

        # Creates a dictionary for each site from quip, calculate its rates for each type of incident and stores it as dictionary inside a list
        list_sites_dictionaries_wbr = []
        for row in reader:
            if row != []:
                site_dict_quip = {'site':f'{row[0].lower()}','week': f'{header_row[0].lower()}', 'worked_hours': float(format(float(row[1]), '.2f'))}
                index = 2
                for incident in list_incidents_counters:
                    site_dict_quip[f'{incident}'] = int(row[index])
                    site_dict_quip[f'{incident}_rate'] = float(format(((int(row[index])/float(row[1]))*200000), '.2f'))
                    index += 1
                del site_dict_quip['events_rate']
                list_sites_dictionaries_wbr.append(site_dict_quip)
            else:
                continue
                
    # Identify wich category is each site loaded from quip, using data from 'Sites and categories.csv'
    for site_dictionary_wbr in list_sites_dictionaries_wbr:
        for site_category in list_sites_categories:
            if site_dictionary_wbr['site'] == site_category['site']:
                site_dictionary_wbr['category'] = site_category['category']
            else:
                continue

    # Creates a list of dictionaries of each category site with its complety data of that week
    # Access each category dictonary, then each site dictionary
    # Matches category with site category
    # Assigne the week value from site dictonary to category dictonary
    # Sum worked hours from each site matched
    # Loop through each key of each category dictonary and site dictionary to match with a incident in a incident list
    # Matches incidents keys of categories dictionaries and sites dictioanries
    # Assigne a value to a key of the category dictionary, adding a value of each site dictioanry
    # Then calculate rate for each key, using a copy of the category site
    for site_category_dictionary_wbr in list_site_categories_dictionaries_wbr:
        for site_dictionary_wbr in list_sites_dictionaries_wbr:
            site_category_dictionary_wbr['week'] = site_dictionary_wbr['week']
            if site_category_dictionary_wbr['category'] == site_dictionary_wbr['category']:
                site_category_dictionary_wbr['worked_hours'] += float(format(site_dictionary_wbr['worked_hours'], '.2f'))
                for key_category in site_category_dictionary_wbr.keys():
                    for key_site in site_dictionary_wbr.keys():
                        for incident in list_incidents_counters:
                            if (key_category == incident) and (key_site == incident):
                                site_category_dictionary_wbr[f'{key_category}'] += site_dictionary_wbr[f'{key_site}']
                            else:
                                continue
            else:
                continue
        # Calculate each incident rate
        for key_category in site_category_dictionary_wbr.copy().keys():
            if key_category not in ('category', 'week', 'worked_hours', 'events'):
                if site_category_dictionary_wbr['worked_hours'] != 0:
                    site_category_dictionary_wbr[f'{key_category}_rate'] = float(format((site_category_dictionary_wbr[f'{key_category}']/site_category_dictionary_wbr['worked_hours'])*200000, '.2f'))
                else:
                    site_category_dictionary_wbr[f'{key_category}_rate'] = 0.00
            else:
                continue

    # This point that are 2 list of dictionaries: a list of sites and a list of categories

    # Get 'Site metrics by week' worksheet sites names again in case a new site was add via quip
    worksheet_site_metrics_by_week_sites_names = get_headers(worksheet_site_metrics_by_week['B'])

    # Writes data on .xlsx data of each category on 'Site metrics by week' and 'Category metrics by week'
    # First, gets each site its column and row indexes
    # Then, assigne a index to each data type present on a list of data from Incidents count and rates.csv
    # Doins so it leaving aside others datas as week and site name
    # Finally, writes on .xlsx all incidents data
    for site_dictionary_wbr in list_sites_dictionaries_wbr:

        site_week_info_col = get_index(worksheet_site_metrics_by_week_headers, site_dictionary_wbr['week']) + 4
        first_site_info_row = get_index(worksheet_site_metrics_by_week_sites_names, site_dictionary_wbr['site']) + 2
        
        list_dicitonaries_data_indexes = []

        for site_data in list_site_data:
            for key, value in site_dictionary_wbr.items():
                if f'{key}' == site_data:
                    data_info_row = list_site_data.index(site_data) + first_site_info_row
                    data_dictionary = {'data': f'{key}', 'index':data_info_row}
                    if data_dictionary not in list_dicitonaries_data_indexes:
                        list_dicitonaries_data_indexes.append(data_dictionary)
                    else:
                        continue
                else:
                    continue

        for key, value in site_dictionary_wbr.items():
            for dictionary_data_index in list_dicitonaries_data_indexes:
                if key == dictionary_data_index['data']:
                    worksheet_site_metrics_by_week.cell(row = dictionary_data_index['index'], column = site_week_info_col).value = value
                else:
                    continue

    for site_category_dictionary_wbr in list_site_categories_dictionaries_wbr:

        categor_week_info_col = get_index(worksheet_category_metrics_by_week_headers, site_category_dictionary_wbr['week']) + 3
        first_category_info_row = get_index(worksheet_category_metrics_by_week_categories_names, site_category_dictionary_wbr['category']) + 2

        list_dicitonaries_data_indexes = []

        for site_data in list_site_data:
            for key, value in site_category_dictionary_wbr.items():
                if f'{key}' == site_data:
                    data_info_row = list_site_data.index(site_data) + first_category_info_row
                    data_dictionary = {'data': f'{key}', 'index':data_info_row}
                    if data_dictionary not in list_dicitonaries_data_indexes:
                        list_dicitonaries_data_indexes.append(data_dictionary)
                    else:
                        continue
                else:
                    continue

        for key, value in site_category_dictionary_wbr.items():
            for dictionary_data_index in list_dicitonaries_data_indexes:
                if key == dictionary_data_index['data']:
                    worksheet_category_metrics_by_week.cell(row = dictionary_data_index['index'], column = categor_week_info_col).value = value
                else:
                    continue
     
    workbook.save(whs_anual_metrics_xlsx)
    workbook.close()