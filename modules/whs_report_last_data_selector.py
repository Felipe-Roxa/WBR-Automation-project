import os
import csv
import openpyxl
from modules.helpers.get_index_with_none_helper import get_index_with_none
from modules.helpers.get_headers_with_none_helper import get_headers_with_none

def last_data_selector(whs_anual_metrics_xlsx, weeks_and_months_csv, quip_directory):
    """This function load a .csv as parameter to input data generated in previous methods"""

    list_files = os.listdir(quip_directory)

    last_item = list_files.pop()
    lenght_last_item = len(last_item)
    week = last_item[lenght_last_item - 6:lenght_last_item - 4]
    current_week = f'wk{week}'
    current_month = ''
    current_quarter = ''
    # Load weeks from Weeks and months.csv
    with open(weeks_and_months_csv) as f:
        reader = csv.reader(f, delimiter = ';')
        weeks_and_months_header_row = next(reader)

        for row in reader:
            if row != [] and (row[0] == current_week):
                current_month = row[1]
                current_quarter = row[2]
            else: 
                continue

        current_year = weeks_and_months_header_row[3]

    current_month = f'{current_month.title()}/{current_year}'
    current_quarter = f'{current_quarter.title()}/{current_year}'
    current_year = f'Year{current_year}'

    # Start openpyxl library to open existing .xlsx file
    workbook = openpyxl.load_workbook(whs_anual_metrics_xlsx)

    # Works on Last site metrics worksheet
    worksheet_last_site_metrics = workbook['Last site metrics']

    # Works on Site metrics by year worksheet
    worksheet_site_metrics_by_year = workbook['Site metrics by year']

    # Works on Site metrics by week worksheet
    worksheet_site_metrics_by_week = workbook['Site metrics by week']

    # Works on Last category metrics worksheet
    worksheet_last_category_metrics = workbook['Last category metrics']

    # Works on Category metrics by year worksheet
    worksheet_category_metrics_by_year = workbook['Category metrics by year']

    # Works on Category metrics by week worksheet
    worksheet_category_metrics_by_week = workbook['Category metrics by week']

    # Fill data in to Last site metrics
    # Get headers from Site metrics by year
    headers_worksheet_site_metrics_by_week_list = get_headers_with_none(worksheet_site_metrics_by_week['1'])

    # Get headers from Site metrics by week
    headers_worksheet_site_metrics_by_year_list = get_headers_with_none(worksheet_site_metrics_by_year['1'])

    current_week_site_index = get_index_with_none(headers_worksheet_site_metrics_by_week_list, current_week) + 1
    current_month_site_index = get_index_with_none(headers_worksheet_site_metrics_by_year_list, current_month) + 1
    current_quarter_site_index = get_index_with_none(headers_worksheet_site_metrics_by_year_list, current_quarter) + 1
    current_year_site_index = get_index_with_none(headers_worksheet_site_metrics_by_year_list, current_year) + 1
    last_week_site_index = current_week_site_index - 1
    last_month_site_index = current_month_site_index - 1

    list_site_columns = []
    list_site_columns.append(worksheet_site_metrics_by_week[openpyxl.utils.cell.get_column_letter(last_week_site_index)])
    list_site_columns.append(worksheet_site_metrics_by_week[openpyxl.utils.cell.get_column_letter(current_week_site_index)])
    list_site_columns.append(worksheet_site_metrics_by_year[openpyxl.utils.cell.get_column_letter(last_month_site_index)])
    list_site_columns.append(worksheet_site_metrics_by_year[openpyxl.utils.cell.get_column_letter(current_month_site_index)])
    list_site_columns.append(worksheet_site_metrics_by_year[openpyxl.utils.cell.get_column_letter(current_quarter_site_index)])
    list_site_columns.append(worksheet_site_metrics_by_year[openpyxl.utils.cell.get_column_letter(current_year_site_index)])

    row = 1
    col = 4
    for column in list_site_columns:
        for data in column:
            worksheet_last_site_metrics.cell(row = row, column = col).value = data.value
            row += 1
        row = 1
        col += 1

    # Fill data in to Last category metrics
    # Get headers from Category metrics by year
    headers_worksheet_category_metrics_by_week_list = get_headers_with_none(worksheet_category_metrics_by_week['1'])

    # Get headers from Category metrics by week
    headers_worksheet_category_metrics_by_year_list = get_headers_with_none(worksheet_category_metrics_by_year['1'])

    current_week_category_index = get_index_with_none(headers_worksheet_category_metrics_by_week_list, current_week) + 1
    current_month_category_index = get_index_with_none(headers_worksheet_category_metrics_by_year_list, current_month) + 1
    current_quarter_category_index = get_index_with_none(headers_worksheet_category_metrics_by_year_list, current_quarter) + 1
    current_year_category_index = get_index_with_none(headers_worksheet_category_metrics_by_year_list, current_year) + 1
    last_week_category_index = current_week_category_index - 1
    last_month_category_index = current_month_category_index - 1

    list_category_columns = []
    list_category_columns.append(worksheet_category_metrics_by_week[openpyxl.utils.cell.get_column_letter(last_week_category_index)])
    list_category_columns.append(worksheet_category_metrics_by_week[openpyxl.utils.cell.get_column_letter(current_week_category_index)])
    list_category_columns.append(worksheet_category_metrics_by_year[openpyxl.utils.cell.get_column_letter(last_month_category_index)])
    list_category_columns.append(worksheet_category_metrics_by_year[openpyxl.utils.cell.get_column_letter(current_month_category_index)])
    list_category_columns.append(worksheet_category_metrics_by_year[openpyxl.utils.cell.get_column_letter(current_quarter_category_index)])
    list_category_columns.append(worksheet_category_metrics_by_year[openpyxl.utils.cell.get_column_letter(current_year_category_index)])

    row = 1
    col = 3
    for column in list_category_columns:
        for data in column:
            worksheet_last_category_metrics.cell(row = row, column = col).value = data.value
            row += 1
        row = 1
        col += 1

    workbook.save(whs_anual_metrics_xlsx)
    workbook.close()