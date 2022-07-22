import openpyxl
from modules.helpers.get_headers_helper import get_headers
from modules.helpers.get_index_helper import get_index
from modules.helpers.ranges_calculator_helper import ranges_calculator
from modules.helpers.get_csv_data_site_helper import get_csv_data_site
from modules.helpers.get_csv_incidents_counters_helper import get_csv_incidents_counters
from modules.helpers.get_csv_incidents_counters_and_hours_helper import get_csv_incidents_counters_and_hours

def site_and_category_calculator_by_month_to_year(whs_anual_metrics_xlsx, weeks_and_months_csv, incidents_count_and_rates_csv):
    """This function manipulates data from a .xlsx file to genreate new data on it"""

    # Load site data from Incidents count and rates.csv to formule 2 lists with keys as variables to work with
    # A list with counters incidents and a list with all data site
    list_incidents_counters = get_csv_incidents_counters(incidents_count_and_rates_csv)
    list_data = get_csv_data_site(incidents_count_and_rates_csv)
    list_incidents_counters_and_hours = get_csv_incidents_counters_and_hours(incidents_count_and_rates_csv)

    # Start openpyxl library to open existing .xlsx file
    workbook = openpyxl.load_workbook(whs_anual_metrics_xlsx)

    # Works on Site metrics by year worksheet
    worksheet_site_metrics_by_year = workbook['Site metrics by year']

    # Works on Site metrics by week worksheet
    worksheet_site_metrics_by_week = workbook['Site metrics by week']

    # Works on Category metrics by year worksheet
    worksheet_category_metrics_by_year = workbook['Category metrics by year']

    # Works on Category metrics by week worksheet
    worksheet_category_metrics_by_week = workbook['Category metrics by week']

    # Get Site metrics by week worksheet site names
    worksheet_site_metrics_by_week_sites_names = get_headers(worksheet_site_metrics_by_week['B'])

    # Get Category metrics by week worksheet site names
    worksheet_category_metrics_by_week_categories_names = get_headers(worksheet_category_metrics_by_week['A'])
    
    # Formule a list of dicitionaries of each site and setting its name to hold information
    list_sites_wbr = set(worksheet_site_metrics_by_week_sites_names)
    list_sites_dictionaries_wbr = []
    for site_wbr in list_sites_wbr:
        site_dictionary_wbr = {'site_name':f'{site_wbr}', 'worked_hours': 0.0}
        for site_data in list_data:
            site_dictionary_wbr[site_data] = 0
        list_sites_dictionaries_wbr.append(site_dictionary_wbr)

    list_site_weeks_ranges = ranges_calculator(weeks_and_months_csv, whs_anual_metrics_xlsx, 'Site metrics by year', 4)

    # Formule a list of dicitionaries of each category and setting its name to hold information
    list_categories_wbr = set(worksheet_category_metrics_by_week_categories_names)
    list_categories_dictionaries_wbr = []
    for category_wbr in list_categories_wbr:
        category_dictionary_wbr = {'category_name':f'{category_wbr}', 'worked_hours': 0.0}
        for category_data in list_data:
            category_dictionary_wbr[category_data] = 0
        list_categories_dictionaries_wbr.append(category_dictionary_wbr)

    list_category_weeks_ranges = ranges_calculator(weeks_and_months_csv, whs_anual_metrics_xlsx, 'Category metrics by year', 3)

    # Works on each anual category dictionary to sum and calculate data yearly
    for category_dictionary_wbr in list_categories_dictionaries_wbr:

        # Get category name index from Category metrics by week worksheet
        first_category_info_row = get_index(worksheet_category_metrics_by_week_categories_names, category_dictionary_wbr['category_name']) + 2

        list_data_index_dictionaries = []

        # Attributes category data and data index relative to category name index, making a dictionary to each data
        for category_data in list_data:
            for key, value in category_dictionary_wbr.items():
                if f'{key}' == category_data:
                    data_info_row = list_data.index(category_data) + first_category_info_row
                    data_index_dictionary = {'data': f'{key}', 'index':data_info_row}
                    if data_index_dictionary not in list_data_index_dictionaries:
                        list_data_index_dictionaries.append(data_index_dictionary)
                    else:
                        continue
                else:
                    continue

        # Remove untreatable information from each anual category dictionary, to sum counts of each type o incident
        category_name = category_dictionary_wbr['category_name']
        del category_dictionary_wbr['category_name']
        
        # Runs through all yearly ranges, select each incident, its index, and filter only the incidents counters to sum in that given range
        for week_range in list_category_weeks_ranges:
            first_index = week_range['first_index']
            last_index = week_range['last_index']
            data_column = week_range['data_column']
            for key, value in category_dictionary_wbr.items():
                for data_index_dictionary in list_data_index_dictionaries:
                    if (key == data_index_dictionary['data']) and (key in list_incidents_counters_and_hours):
                        incident_index = data_index_dictionary['index']
                        for column in range (first_index, last_index):
                            cell_value = worksheet_category_metrics_by_week.cell(row = incident_index, column = column).value
                            if cell_value != None:
                                category_dictionary_wbr[f'{key}'] += cell_value
                            else:
                                continue
                    else:
                        continue

            # Makes a dictionary copy to calculate rates from each incident in category dictionary, using string to match the rate with it correponding counter
            for key in category_dictionary_wbr.copy().keys():
                if key in list_incidents_counters:
                    if category_dictionary_wbr['worked_hours'] != 0:
                        category_dictionary_wbr[f'{key}_rate'] = float(format(((category_dictionary_wbr[f'{key}']/category_dictionary_wbr['worked_hours'])*200000), '.2f'))
                    else:
                        category_dictionary_wbr[f'{key}'] = 0
                else:
                    continue

            # Writes on yearly worksheet
            for key, value in category_dictionary_wbr.items():
                for data_index_dictionary in list_data_index_dictionaries:
                    if key == data_index_dictionary['data']:
                        worksheet_category_metrics_by_year.cell(row = data_index_dictionary['index'], column = data_column).value = value
                    else:
                        continue
            
            # Erases data from category dictionary to not populate after current date
            for key, value in category_dictionary_wbr.items():
                category_dictionary_wbr[f'{key}'] = 0
        
        # Put information back to category dictionary
        category_dictionary_wbr['category_name'] = category_name

    # Works on each anual site dictionary to sum and calculate data yearly
    for site_dictionary_wbr in list_sites_dictionaries_wbr:

        # Get site name index from Site metrics by week worksheet
        first_site_info_row = get_index(worksheet_site_metrics_by_week_sites_names, site_dictionary_wbr['site_name']) + 2

        list_data_index_dictionaries = []

        # Attributes site data and data index relative to site name index, making a dictionary to each data
        for site_data in list_data:
            for key, value in site_dictionary_wbr.items():
                if f'{key}' == site_data:
                    data_info_row = list_data.index(site_data) + first_site_info_row
                    data_index_dictionary = {'data': f'{key}', 'index':data_info_row}
                    if data_index_dictionary not in list_data_index_dictionaries:
                        list_data_index_dictionaries.append(data_index_dictionary)
                    else:
                        continue
                else:
                    continue

        # Remove untreatable information from each anual site dictionary, to sum counts of each type o incident
        site_name = site_dictionary_wbr['site_name']
        del site_dictionary_wbr['site_name']
        
        # Runs through all yearly ranges, select each incident, its index, and filter only the incidents counters to sum in that given range
        for week_range in list_site_weeks_ranges:
            first_index = week_range['first_index']
            last_index = week_range['last_index']
            data_column = week_range['data_column']
            for key, value in site_dictionary_wbr.items():
                for data_index_dictionary in list_data_index_dictionaries:
                    if (key == data_index_dictionary['data']) and (key in list_incidents_counters_and_hours):
                        incident_index = data_index_dictionary['index']
                        for column in range (first_index, last_index):
                            cell_value = worksheet_site_metrics_by_week.cell(row = incident_index, column = column).value
                            if cell_value != None:
                                site_dictionary_wbr[f'{key}'] += cell_value
                            else:
                                continue
                    else:
                        continue

            # Makes a dictionary copy to calculate rates from each incident in site dictionary, using string to match the rate with it correponding counter
            for key in site_dictionary_wbr.copy().keys():
                if key in list_incidents_counters:
                    if site_dictionary_wbr['worked_hours'] != 0:
                        site_dictionary_wbr[f'{key}_rate'] = float(format(((site_dictionary_wbr[f'{key}']/site_dictionary_wbr['worked_hours'])*200000), '.2f'))
                    else:
                        site_dictionary_wbr[f'{key}'] = 0
                else:
                    continue

            # Writes on yearly worksheet
            for key, value in site_dictionary_wbr.items():
                for data_index_dictionary in list_data_index_dictionaries:
                    if key == data_index_dictionary['data']:
                        worksheet_site_metrics_by_year.cell(row = data_index_dictionary['index'], column = data_column).value = value
                    else:
                        continue
            
            # Erases data from site dictionary to not populate after current date
            for key, value in site_dictionary_wbr.items():
                site_dictionary_wbr[f'{key}'] = 0
        
        # Put information back to site dictionary
        site_dictionary_wbr['site_name'] = site_name

    workbook.save(whs_anual_metrics_xlsx)
    workbook.close()